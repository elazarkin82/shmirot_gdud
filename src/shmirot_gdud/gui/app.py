import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import json
from typing import List, Dict, Optional
import pandas as pd
from openpyxl.styles import PatternFill, Alignment
from datetime import datetime, timedelta

from shmirot_gdud.core.models import Group, Schedule, ScheduleRange, ScheduleSlot
from shmirot_gdud.core.scheduler import Scheduler
from shmirot_gdud.core.config import config
from shmirot_gdud.core.constraints.factory import ConstraintFactory
from shmirot_gdud.core.constraints.implementations import UnavailabilityConstraint, ActivityWindowConstraint, DateSpecificConstraint, StaffingRuleConstraint
from shmirot_gdud.gui.dialogs import GroupCreationDialog, DateRangeDialog, ImprovementSettingsDialog, AdvancedSettingsDialog, StaffingExceptionsDialog
from shmirot_gdud.gui.schedule_grid import ScheduleGrid, DISABLED_ID
from shmirot_gdud.gui.utils import bidi_text

class App:
    def __init__(self, root):
        self.root = root
        self.root.title(bidi_text("מערכת שיבוץ שמירות גדודית"))
        self.root.geometry("1400x800")

        config.save()

        self.groups: List[Group] = []
        self.schedule: Optional[Schedule] = None

        self._create_menu()
        self._show_main_menu()

    def _create_menu(self):
        menubar = tk.Menu(self.root)
        
        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label=bidi_text("טען קבוצות"), command=self._load_groups)
        file_menu.add_command(label=bidi_text("שמור קבוצות"), command=self._save_groups)
        file_menu.add_separator()
        file_menu.add_command(label=bidi_text("טען סידור עבודה"), command=self._load_schedule)
        file_menu.add_command(label=bidi_text("שמור סידור עבודה"), command=self._save_schedule)
        file_menu.add_separator()
        file_menu.add_command(label=bidi_text("ייצוא לאקסל"), command=self._export_excel)
        file_menu.add_separator()
        file_menu.add_command(label=bidi_text("הגדרות ניקוד"), command=self._open_advanced_settings)
        file_menu.add_separator()
        file_menu.add_command(label=bidi_text("יציאה"), command=self.root.quit)
        menubar.add_cascade(label=bidi_text("קובץ"), menu=file_menu)
        
        nav_menu = tk.Menu(menubar, tearoff=0)
        nav_menu.add_command(label=bidi_text("תפריט ראשי"), command=self._show_main_menu)
        nav_menu.add_command(label=bidi_text("ניהול קבוצות"), command=self._show_group_management)
        nav_menu.add_command(label=bidi_text("לוח שיבוץ"), command=self._show_schedule)
        menubar.add_cascade(label=bidi_text("ניווט"), menu=nav_menu)
        
        self.root.config(menu=menubar)

    def _clear_window(self):
        for widget in self.root.winfo_children():
            if isinstance(widget, tk.Menu): continue
            widget.destroy()

    def _show_main_menu(self):
        self._clear_window()
        
        frame = ttk.Frame(self.root, padding=50)
        frame.pack(expand=True)
        
        ttk.Label(frame, text=bidi_text("מערכת שיבוץ שמירות"), font=("Arial", 24, "bold")).pack(pady=30)
        
        btn_width = 25
        ttk.Button(frame, text=bidi_text("יצירת קבוצה חדשה"), width=btn_width, command=self._open_create_group_dialog).pack(pady=10)
        ttk.Button(frame, text=bidi_text("ניהול קבוצות"), width=btn_width, command=self._show_group_management).pack(pady=10)
        ttk.Button(frame, text=bidi_text("יצירת/צפייה בלוח שיבוץ"), width=btn_width, command=self._show_schedule).pack(pady=10)
        ttk.Button(frame, text=bidi_text("שמירת נתונים"), width=btn_width, command=self._save_groups).pack(pady=10)
        ttk.Button(frame, text=bidi_text("טעינת נתונים"), width=btn_width, command=self._load_groups).pack(pady=10)

    def _open_create_group_dialog(self):
        def on_create(group: Group):
            max_id = 0
            for g in self.groups:
                try:
                    gid = int(g.id)
                    if gid > max_id: max_id = gid
                except ValueError: pass
            group.id = str(max_id + 1)
            self.groups.append(group)
            messagebox.showinfo(bidi_text("הצלחה"), bidi_text(f"הקבוצה {group.name} נוצרה בהצלחה"))
            
        GroupCreationDialog(self.root, on_create)

    def _show_group_management(self):
        self._clear_window()
        
        main_paned = ttk.PanedWindow(self.root, orient=tk.HORIZONTAL)
        main_paned.pack(fill=tk.BOTH, expand=True)

        left_frame = ttk.Frame(main_paned, width=300)
        main_paned.add(left_frame, weight=1)
        
        ttk.Label(left_frame, text=bidi_text("רשימת קבוצות"), font=("Arial", 14, "bold")).pack(pady=5)
        
        self.group_listbox = tk.Listbox(left_frame, height=20, justify="right")
        self.group_listbox.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.group_listbox.bind('<<ListboxSelect>>', self._on_group_select)
        
        btn_frame = ttk.Frame(left_frame)
        btn_frame.pack(fill=tk.X, padx=5, pady=5)
        ttk.Button(btn_frame, text=bidi_text("הוסף קבוצה"), command=self._open_create_group_dialog_refresh).pack(side=tk.RIGHT, fill=tk.X, expand=True)
        ttk.Button(btn_frame, text=bidi_text("מחק קבוצה"), command=self._delete_group).pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        right_frame = ttk.Frame(main_paned)
        main_paned.add(right_frame, weight=3)
        
        self.details_frame = ttk.LabelFrame(right_frame, text=bidi_text("פרטי קבוצה"))
        self.details_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        info_frame = ttk.Frame(self.details_frame)
        info_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Label(info_frame, text=bidi_text("שם הקבוצה:")).grid(row=0, column=2, sticky=tk.E, padx=5)
        self.name_var = tk.StringVar()
        ttk.Entry(info_frame, textvariable=self.name_var, justify="right").grid(row=0, column=1, sticky=tk.EW)
        
        ttk.Label(info_frame, text=bidi_text("גודל סד\"כ:")).grid(row=1, column=2, sticky=tk.E, padx=5)
        self.staffing_var = tk.StringVar()
        ttk.Entry(info_frame, textvariable=self.staffing_var, justify="right").grid(row=1, column=1, sticky=tk.EW)
        
        ttk.Label(info_frame, text=bidi_text("מכסה שבועית (קשיח):")).grid(row=2, column=2, sticky=tk.E, padx=5)
        self.quota_var = tk.StringVar()
        ttk.Entry(info_frame, textvariable=self.quota_var, justify="right").grid(row=2, column=1, sticky=tk.EW)
        
        self.simultaneous_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(info_frame, text=bidi_text("מאפשר שמירה כפולה (בו-זמנית)"), variable=self.simultaneous_var).grid(row=3, column=0, columnspan=3, sticky=tk.E, pady=5)
        
        info_frame.columnconfigure(1, weight=1)

        # Constraints Buttons
        constraints_frame = ttk.Frame(self.details_frame)
        constraints_frame.pack(fill=tk.X, padx=5, pady=10)
        
        self._create_constraint_button(constraints_frame, "ניהול אי-זמינות", UnavailabilityConstraint)
        self._create_constraint_button(constraints_frame, "ניהול חלונות פעילות", ActivityWindowConstraint)
        self._create_constraint_button(constraints_frame, "ניהול אילוצי תאריכים", DateSpecificConstraint)
        self._create_constraint_button(constraints_frame, "ניהול חוקי איוש", StaffingRuleConstraint)
        
        ttk.Button(constraints_frame, text=bidi_text("ניהול חריגות סד\"כ"), command=self._manage_staffing_exceptions).pack(fill=tk.X, pady=2)
        
        ttk.Button(self.details_frame, text=bidi_text("שמור שינויים"), command=self._save_group_details).pack(fill=tk.X, padx=5, pady=20)

        self._refresh_group_list()

    def _create_constraint_button(self, parent, text, constraint_class):
        def command():
            self._manage_constraint(constraint_class)
        
        frame = ttk.Frame(parent)
        frame.pack(fill=tk.X, pady=2)
        
        btn = ttk.Button(frame, text=bidi_text(text), command=command)
        btn.pack(side=tk.RIGHT, fill=tk.X, expand=True)
        
        # Status label
        lbl = ttk.Label(frame, text="", width=15, anchor="e")
        lbl.pack(side=tk.LEFT, padx=5)
        
        # Store reference to label to update it later
        if not hasattr(self, 'constraint_labels'):
            self.constraint_labels = {}
        self.constraint_labels[constraint_class] = lbl

    def _manage_constraint(self, constraint_class):
        selection = self.group_listbox.curselection()
        if not selection: return
        idx = selection[0]
        group = self.groups[idx]
        
        # Find existing constraint or create new
        constraint = next((c for c in group.constraints if isinstance(c, constraint_class)), None)
        if not constraint:
            constraint = constraint_class()
            group.constraints.append(constraint)
            
        def on_save(updated_constraint):
            # Update status label
            if constraint_class in self.constraint_labels:
                self.constraint_labels[constraint_class].config(text=bidi_text(updated_constraint.get_status_text()))
        
        constraint.open_edit_dialog(self.root, on_save)

    def _manage_staffing_exceptions(self):
        selection = self.group_listbox.curselection()
        if not selection: return
        idx = selection[0]
        group = self.groups[idx]
        
        def on_save(exceptions):
            group.staffing_exceptions = exceptions
            
        StaffingExceptionsDialog(self.root, bidi_text(f"חריגות סד\"כ עבור {group.name}"), group.staffing_exceptions, on_save)

    def _open_create_group_dialog_refresh(self):
        def on_create(group: Group):
            max_id = 0
            for g in self.groups:
                try:
                    gid = int(g.id)
                    if gid > max_id: max_id = gid
                except ValueError: pass
            group.id = str(max_id + 1)
            self.groups.append(group)
            self._refresh_group_list()
        GroupCreationDialog(self.root, on_create)

    def _show_schedule(self):
        self._clear_window()
        
        main_paned = ttk.PanedWindow(self.root, orient=tk.HORIZONTAL)
        main_paned.pack(fill=tk.BOTH, expand=True)

        stats_frame = ttk.Frame(main_paned, width=350)
        main_paned.add(stats_frame, weight=1)
        
        ttk.Label(stats_frame, text=bidi_text("סטטיסטיקות שיבוץ"), font=("Arial", 14, "bold")).pack(pady=10)
        
        self.stats_tree = ttk.Treeview(stats_frame, columns=("Name", "Staffing", "Count", "Percent", "Hard"), show="headings")
        self.stats_tree.heading("Name", text=bidi_text("קבוצה"))
        self.stats_tree.heading("Staffing", text=bidi_text("סד\"כ"))
        self.stats_tree.heading("Count", text=bidi_text("משמרות"))
        self.stats_tree.heading("Percent", text=bidi_text("%"))
        self.stats_tree.heading("Hard", text=bidi_text("לילה (2-6)"))
        
        self.stats_tree.column("Name", width=90, anchor="center")
        self.stats_tree.column("Staffing", width=40, anchor="center")
        self.stats_tree.column("Count", width=50, anchor="center")
        self.stats_tree.column("Percent", width=40, anchor="center")
        self.stats_tree.column("Hard", width=70, anchor="center")
        
        self.stats_tree.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.stats_tree.bind('<<TreeviewSelect>>', self._on_stats_select)

        ttk.Button(stats_frame, text=bidi_text("בטל הדגשה"), command=self._clear_highlight).pack(fill=tk.X, padx=5, pady=5)

        right_container = ttk.Frame(main_paned)
        main_paned.add(right_container, weight=4)
        
        top_frame = ttk.Frame(right_container, padding=10)
        top_frame.pack(fill=tk.X)
        
        ttk.Button(top_frame, text=bidi_text("חזור לתפריט ראשי"), command=self._show_main_menu).pack(side=tk.RIGHT)
        
        ttk.Button(top_frame, text=bidi_text("1. צור טבלה ריקה"), command=self._open_date_range_dialog).pack(side=tk.LEFT, padx=5)
        
        self.fill_btn = ttk.Button(top_frame, text=bidi_text("2. מלא אוטומטית"), command=self._fill_schedule, state="disabled")
        self.fill_btn.pack(side=tk.LEFT, padx=5)
        
        self.improve_btn = ttk.Button(top_frame, text=bidi_text("3. שפר סידור"), command=self._open_improvement_dialog, state="disabled")
        self.improve_btn.pack(side=tk.LEFT, padx=5)
        
        ttk.Button(top_frame, text=bidi_text("ייצוא לאקסל"), command=self._export_excel).pack(side=tk.LEFT, padx=5)

        zoom_frame = ttk.Frame(top_frame)
        zoom_frame.pack(side=tk.LEFT, padx=20)
        ttk.Label(zoom_frame, text=bidi_text("זום:")).pack(side=tk.LEFT)
        ttk.Button(zoom_frame, text="-", width=3, command=lambda: self.schedule_grid.zoom(-0.1)).pack(side=tk.LEFT)
        ttk.Button(zoom_frame, text="+", width=3, command=lambda: self.schedule_grid.zoom(0.1)).pack(side=tk.LEFT)
        ttk.Button(zoom_frame, text=bidi_text("התאם לרוחב"), command=lambda: self.schedule_grid.fit_to_width()).pack(side=tk.LEFT, padx=5)

        grid_frame = ttk.Frame(right_container)
        grid_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        self.schedule_grid = ScheduleGrid(grid_frame, self.groups, self._on_schedule_change, bg="white")
        
        h_scroll = ttk.Scrollbar(grid_frame, orient=tk.HORIZONTAL, command=self.schedule_grid.xview)
        v_scroll = ttk.Scrollbar(grid_frame, orient=tk.VERTICAL, command=self.schedule_grid.yview)
        
        self.schedule_grid.configure(xscrollcommand=h_scroll.set, yscrollcommand=v_scroll.set)
        
        h_scroll.pack(side=tk.BOTTOM, fill=tk.X)
        v_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.schedule_grid.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        if self.schedule:
            self.schedule_grid.set_schedule(self.schedule)
            self._update_stats_content()
            self.fill_btn.configure(state="normal")
            self.improve_btn.configure(state="normal")

    def _on_stats_select(self, event):
        selection = self.stats_tree.selection()
        if not selection: return
        item = selection[0]
        values = self.stats_tree.item(item, "values")
        if not values: return
        group_name_bidi = values[0]
        found_id = None
        for g in self.groups:
            if bidi_text(g.name) == group_name_bidi:
                found_id = g.id
                break
        self.schedule_grid.set_highlighted_group(found_id)

    def _clear_highlight(self):
        if hasattr(self, 'stats_tree'):
            for item in self.stats_tree.selection():
                self.stats_tree.selection_remove(item)
        if hasattr(self, 'schedule_grid'):
            self.schedule_grid.set_highlighted_group(None)

    def _open_date_range_dialog(self):
        def on_confirm(start_date: str, end_date: str):
            self.schedule = Schedule.create_empty(start_date, end_date)
            self.schedule_grid.set_schedule(self.schedule)
            self._update_stats_content()
            self.fill_btn.configure(state="normal")
            self.improve_btn.configure(state="normal")
        DateRangeDialog(self.root, on_confirm)

    def _fill_schedule(self):
        if not self.schedule or not self.groups: return
        self.root.config(cursor="watch")
        self.root.update()
        try:
            scheduler = Scheduler(self.groups)
            self.schedule = scheduler.fill_schedule(self.schedule)
            self.schedule_grid.set_schedule(self.schedule)
            self._update_stats_content()
            messagebox.showinfo(bidi_text("הצלחה"), bidi_text("השיבוץ הושלם"))
        finally:
            self.root.config(cursor="")

    def _open_improvement_dialog(self):
        if not self.schedule: return
        def on_confirm(hard_start: int, hard_end: int):
            self._improve_current_schedule(hard_start, hard_end)
        ImprovementSettingsDialog(self.root, on_confirm)

    def _open_advanced_settings(self):
        AdvancedSettingsDialog(self.root)

    def _improve_current_schedule(self, hard_start: int = 2, hard_end: int = 6):
        if not self.schedule: return
        progress_win = tk.Toplevel(self.root)
        progress_win.title(bidi_text("משפר סידור..."))
        progress_win.geometry("350x150")
        progress_win.transient(self.root)
        progress_win.grab_set()
        
        ttk.Label(progress_win, text=bidi_text("אנא המתן, מבצע אופטימיזציה...")).pack(pady=10)
        progress_bar = ttk.Progressbar(progress_win, orient=tk.HORIZONTAL, length=300, mode='determinate')
        progress_bar.pack(pady=5)
        percent_label = ttk.Label(progress_win, text="0.00%")
        percent_label.pack(pady=5)
        self.root.update()
        
        def update_progress(val):
            progress_bar['value'] = val
            percent_label.config(text=f"{val:.2f}%")
            self.root.update()

        try:
            scheduler = Scheduler(self.groups)
            scheduler.schedule = self.schedule
            self.schedule = scheduler.improve_schedule(hard_start, hard_end, update_progress)
            self.schedule_grid.set_schedule(self.schedule)
            self._update_stats_content()
            progress_win.destroy()
            messagebox.showinfo(bidi_text("הצלחה"), bidi_text("השיפור הושלם"))
        except Exception as e:
            progress_win.destroy()
            messagebox.showerror(bidi_text("שגיאה"), str(e))

    def _update_stats_content(self):
        if not self.schedule or not self.groups: return
        selected_group_id = None
        if hasattr(self, 'schedule_grid') and self.schedule_grid.highlighted_group_id:
            selected_group_id = self.schedule_grid.highlighted_group_id

        for item in self.stats_tree.get_children():
            self.stats_tree.delete(item)
            
        valid_slots = [s for s in self.schedule.slots if s.group_id != DISABLED_ID]
        total_slots = len(valid_slots)
        group_counts = {g.id: 0 for g in self.groups}
        group_hard_counts = {g.id: 0 for g in self.groups}
        
        for slot in valid_slots:
            if slot.group_id and slot.group_id in group_counts:
                group_counts[slot.group_id] += 1
                if 2 <= slot.hour < 6:
                    group_hard_counts[slot.group_id] += 1
                
        for g in self.groups:
            count = group_counts[g.id]
            hard_count = group_hard_counts[g.id]
            percent = (count / total_slots * 100) if total_slots > 0 else 0
            hard_percent = (hard_count / count * 100) if count > 0 else 0
            staffing = str(g.staffing_size) if g.staffing_size is not None else "-"
            
            item_id = self.stats_tree.insert("", tk.END, values=(
                bidi_text(g.name), staffing, count, f"{percent:.1f}%", f"{hard_percent:.1f}%"
            ))
            if selected_group_id and g.id == selected_group_id:
                self.stats_tree.selection_set(item_id)

    def _update_stats(self):
        if hasattr(self, 'stats_tree'):
            self._update_stats_content()

    def _delete_group(self):
        selection = self.group_listbox.curselection()
        if selection:
            if messagebox.askyesno(bidi_text("אישור מחיקה"), bidi_text("האם אתה בטוח שברצונך למחוק את הקבוצה?")):
                idx = selection[0]
                del self.groups[idx]
                self._refresh_group_list()
                self._clear_details()

    def _refresh_group_list(self):
        if not hasattr(self, 'group_listbox'): return
        self.group_listbox.delete(0, tk.END)
        for g in self.groups:
            self.group_listbox.insert(tk.END, bidi_text(g.name))

    def _on_group_select(self, event):
        selection = self.group_listbox.curselection()
        if selection:
            idx = selection[0]
            group = self.groups[idx]
            self.name_var.set(group.name)
            self.staffing_var.set(str(group.staffing_size) if group.staffing_size is not None else "")
            self.quota_var.set(str(group.weekly_guard_quota) if group.weekly_guard_quota is not None else "")
            self.simultaneous_var.set(group.can_guard_simultaneously)
            
            # Update status labels for constraints
            if hasattr(self, 'constraint_labels'):
                for cls, lbl in self.constraint_labels.items():
                    constraint = next((c for c in group.constraints if isinstance(c, cls)), None)
                    text = constraint.get_status_text() if constraint else "0 חוקים"
                    lbl.config(text=bidi_text(text))

    def _save_group_details(self):
        selection = self.group_listbox.curselection()
        if selection:
            idx = selection[0]
            group = self.groups[idx]
            group.name = self.name_var.get()
            group.can_guard_simultaneously = self.simultaneous_var.get()
            try:
                s_size = self.staffing_var.get()
                group.staffing_size = int(s_size) if s_size else None
            except ValueError:
                messagebox.showerror(bidi_text("שגיאה"), bidi_text("גודל סד\"כ חייב להיות מספר שלם"))
                return
            try:
                quota = self.quota_var.get()
                group.weekly_guard_quota = int(quota) if quota else None
            except ValueError:
                messagebox.showerror(bidi_text("שגיאה"), bidi_text("מכסה שבועית חייבת להיות מספר שלם"))
                return
            if not group.validate():
                messagebox.showwarning(bidi_text("אזהרה"), bidi_text("לקבוצה חייב להיות מוגדר סד\"כ או מכסה שבועית"))
            self._refresh_group_list()
            self.group_listbox.selection_set(idx)
            messagebox.showinfo(bidi_text("הצלחה"), bidi_text("השינויים נשמרו"))

    def _save_groups(self):
        filename = filedialog.asksaveasfilename(defaultextension=".json", filetypes=[("JSON files", "*.json")])
        if filename:
            data = [g.to_dict() for g in self.groups]
            with open(filename, 'w') as f:
                json.dump(data, f, indent=2)
            messagebox.showinfo(bidi_text("הצלחה"), bidi_text("הקבוצות נשמרו בהצלחה"))

    def _load_groups(self):
        filename = filedialog.askopenfilename(filetypes=[("JSON files", "*.json")])
        if filename:
            try:
                with open(filename, 'r') as f:
                    data = json.load(f)
                self.groups = []
                for d in data:
                    self.groups.append(Group.from_dict(d))
                messagebox.showinfo(bidi_text("הצלחה"), bidi_text("הקבוצות נטענו בהצלחה"))
                if hasattr(self, 'group_listbox'):
                    self._refresh_group_list()
                    self._clear_details()
            except Exception as e:
                messagebox.showerror(bidi_text("שגיאה"), bidi_text(f"נכשל בטעינת הקבוצות: {e}"))

    def _save_schedule(self):
        if not self.schedule:
            messagebox.showwarning(bidi_text("אזהרה"), bidi_text("אין סידור עבודה לשמירה"))
            return
        filename = filedialog.asksaveasfilename(defaultextension=".json", filetypes=[("JSON files", "*.json")])
        if filename:
            try:
                data = {
                    "groups": [g.to_dict() for g in self.groups],
                    "schedule": self.schedule.to_dict()
                }
                with open(filename, 'w') as f:
                    json.dump(data, f, indent=2)
                messagebox.showinfo(bidi_text("הצלחה"), bidi_text("הסידור נשמר בהצלחה"))
            except Exception as e:
                messagebox.showerror(bidi_text("שגיאה"), bidi_text(f"נכשל בשמירת הסידור: {e}"))

    def _load_schedule(self):
        filename = filedialog.askopenfilename(filetypes=[("JSON files", "*.json")])
        if filename:
            try:
                with open(filename, 'r') as f:
                    data = json.load(f)
                self.groups = []
                for d in data.get("groups", []):
                    self.groups.append(Group.from_dict(d))
                if "schedule" in data:
                    self.schedule = Schedule.from_dict(data["schedule"])
                else:
                    self.schedule = None
                messagebox.showinfo(bidi_text("הצלחה"), bidi_text("הסידור נטען בהצלחה"))
                self._show_schedule()
            except Exception as e:
                messagebox.showerror(bidi_text("שגיאה"), bidi_text(f"נכשל בטעינת הסידור: {e}"))

    def _export_excel(self):
        if not self.schedule:
            messagebox.showwarning(bidi_text("אזהרה"), bidi_text("יש לייצר סידור עבודה תחילה"))
            return
        filename = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if filename:
            schedule_data = []
            days_names = ["ראשון", "שני", "שלישי", "רביעי", "חמישי", "שישי", "שבת"]
            start_date = datetime.strptime(self.schedule.start_date, "%Y-%m-%d")
            end_date = datetime.strptime(self.schedule.end_date, "%Y-%m-%d")
            num_days = (end_date - start_date).days + 1
            slot_map = {}
            for slot in self.schedule.slots:
                slot_map[(slot.date, slot.hour, slot.position)] = slot.group_id
            for hour in range(24):
                row = {"שעה": f"{hour:02d}:00 - {hour+1:02d}:00"}
                current = start_date
                for d in range(num_days):
                    date_str = current.strftime("%Y-%m-%d")
                    py_wd = current.weekday()
                    our_wd = (py_wd + 1) % 7
                    day_name = days_names[our_wd]
                    header = f"{day_name} {current.strftime('%d/%m')}"
                    g1_id = slot_map.get((date_str, hour, 1))
                    g2_id = slot_map.get((date_str, hour, 2))
                    def get_name(gid):
                        if gid == DISABLED_ID: return "---"
                        return next((g.name for g in self.groups if g.id == gid), "")
                    g1_name = get_name(g1_id)
                    g2_name = get_name(g2_id)
                    row[f"{header} עמדה 1"] = g1_name
                    row[f"{header} עמדה 2"] = g2_name
                    current += timedelta(days=1)
                schedule_data.append(row)
            df_schedule = pd.DataFrame(schedule_data)
            groups_data = []
            for g in self.groups:
                groups_data.append({
                    "שם": g.name,
                    "סד\"כ": g.staffing_size,
                    "מכסה שבועית": g.weekly_guard_quota,
                    "מאפשר שמירה כפולה": "כן" if g.can_guard_simultaneously else "לא",
                    "אי-זמינות": "; ".join([f"יום {r.day} {r.start_hour}-{r.end_hour}" for r in g.hard_unavailability_rules]),
                    "חלונות פעילות": "; ".join([f"יום {r.day} {r.start_hour}-{r.end_hour}" for r in g.primary_activity_windows])
                })
            df_groups = pd.DataFrame(groups_data)
            valid_slots = [s for s in self.schedule.slots if s.group_id != DISABLED_ID]
            total_slots = len(valid_slots)
            group_counts = {g.id: 0 for g in self.groups}
            for slot in valid_slots:
                if slot.group_id and slot.group_id in group_counts:
                    group_counts[slot.group_id] += 1
            stats_data = []
            for g in self.groups:
                count = group_counts[g.id]
                percent = (count / total_slots * 100) if total_slots > 0 else 0
                stats_data.append({
                    "קבוצה": g.name,
                    "סד\"כ": g.staffing_size,
                    "משמרות": count,
                    "אחוז": f"{percent:.1f}%"
                })
            df_stats = pd.DataFrame(stats_data)
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df_schedule.to_excel(writer, sheet_name='לוח שיבוץ', index=False)
                df_groups.to_excel(writer, sheet_name='קבוצות', index=False)
                df_stats.to_excel(writer, sheet_name='סטטיסטיקות', index=False)
                workbook = writer.book
                ws = workbook['לוח שיבוץ']
                ws.sheet_view.rightToLeft = True
                color_map = {g.name: g.color.replace("#", "") for g in self.groups}
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except: pass
                    adjusted_width = (max_length + 2) * 1.2
                    ws.column_dimensions[column_letter].width = adjusted_width
                for row in ws.iter_rows(min_row=2, min_col=2):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        if cell.value in color_map:
                            fill = PatternFill(start_color=color_map[cell.value], end_color=color_map[cell.value], fill_type="solid")
                            cell.fill = fill
                        elif cell.value == "---":
                            fill = PatternFill(start_color="555555", end_color="555555", fill_type="solid")
                            cell.fill = fill
                workbook['קבוצות'].sheet_view.rightToLeft = True
                workbook['סטטיסטיקות'].sheet_view.rightToLeft = True
                for sheet_name in ['קבוצות', 'סטטיסטיקות']:
                    ws_other = workbook[sheet_name]
                    for column in ws_other.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except: pass
                        adjusted_width = (max_length + 2) * 1.2
                        ws_other.column_dimensions[column_letter].width = adjusted_width
            messagebox.showinfo(bidi_text("הצלחה"), bidi_text("הייצוא הושלם בהצלחה"))

    def _on_schedule_change(self) -> bool:
        # Validate the change
        if not self.schedule: return False
        
        scheduler = Scheduler(self.groups)
        scheduler.schedule = self.schedule
        errors = scheduler.validate_schedule()
        
        if errors:
             msg = "\n".join(errors[:10])
             if len(errors) > 10:
                 msg += "\n..."
             messagebox.showwarning(bidi_text("שגיאה בשיבוץ"), bidi_text(msg))
             return False # Invalid move
        
        self._update_stats()
        return True # Valid move

def main():
    root = tk.Tk()
    app = App(root)
    root.mainloop()
